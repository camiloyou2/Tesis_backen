from fastapi import APIRouter, File, Form , Header ,HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse
from  pydantic import BaseModel , EmailStr
from configdatabase.configdbs import Database
from configdatabase.configdbs import Database
from jwtconfig.functions_jwt import  write_token , validar_token
from configdatabase.configdbs import Database
from models.modelos_estradada_api import User
from openpyxl import load_workbook

from models.models_power import Proyecto # type: ignore
api_power = APIRouter()
from fastapi import FastAPI, File, UploadFile
from openpyxl import load_workbook
import tempfile
from tabulate import tabulate




api_power = APIRouter()

@api_power.post("/load_excel")
async def upload_file(
    excel: UploadFile = File(None)
):
    db = Database()
    excel_content = await excel.read() if excel else None

    # Crear un archivo temporal con la extensión correcta
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(excel_content)
        tmp_path = tmp.name

    # Cargar el archivo de Excel usando openpyxl
    workbook = load_workbook(filename=tmp_path)

    # Iterar a través de cada hoja en el libro de trabajo
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Data from sheet '{sheet_name}'::")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip the header row
            if i == 0:
                continue

            # Process the row as described
            part_document = row[1].split(" -- ")

            id_documento = db.get_tipo_documento(part_document[0])
            id_opcion_de_grado = db.get_tipo_opciones_de_grado(row[4])
            id_estado = db.get_tipo_Estado(row[6])

            # Create the student record in the database
            db.create_estudiante(
                part_document[1],     # numero_documento
                row[2],               # nombre
                row[0],               # codigo
                row[3],               # semestre
                row[7],               # creditos
                id_opcion_de_grado,   # id_opcion
                id_documento,         # id_documento
                id_estado,            # id_estado
                row[5]                # solicitudes_sis
            )

            
        
                

        
    
    
    return {"status": "Success", "message": "Files and form data received"}

@api_power.get("/get_opciones_de_grado")
async def get_opciones_de_grado(
  
): 
    db = Database()

    return db.get_all_opciones_de_grado()

@api_power.get("/get_tipo_documento")
async def get_tipo_documento(
  
): 
    db = Database()

    return db.get_all_tipo_documento()

@api_power.get("/get_all_Estado")
async def get_all_Estado(
  
): 
    db = Database()

    return db.get_all_Estado()

@api_power.post("/load_graduados")
async def load_graduados(
    exceltwo: UploadFile = File(None)
):
    db = Database()
    excel_content = await exceltwo.read() if exceltwo else None

    # Crear un archivo temporal con la extensión correcta
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(excel_content)
        tmp_path = tmp.name

    # Cargar el archivo de Excel usando openpyxl
    workbook = load_workbook(filename=tmp_path)

    # Iterar a través de cada hoja en el libro de trabajo
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Data from sheet '{sheet_name}':")
        for row in sheet.iter_rows(values_only=True):
           
            db.insertar_graduados(row[1],row[2],row[5],f"{row[6]} {row[7]} {row[8]} {row[9]}",row[20],row[21],row[22],row[27] ,row[13])
   
          
                

      
    
    return {"status": "Success", "message": "Files and form data received"}


  # db = Database()
    #db.ceate_pasantia(id_profesor,nombre,anteproyecto_content,acta_content
                   #   ,certificado_laboral_content, certificado_arl_content , 
                     # horario_del_pasante_content, fecha_inicio, terminada)

    # Return response

  #  TRUNCATE `prueba`.`directores`;
#TRUNCATE `prueba`.`actas`;
#TRUNCATE `prueba`.`directores`;
#TRUNCATE `prueba`.`estudiantes`;
#TRUNCATE `prueba`.`periodos`;
#TRUNCATE `prueba`.`modalidades`;
#TRUNCATE `prueba`.`periodos`;
#TRUNCATE `prueba`.`proyectos`;
#TRUNCATE `prueba`.`proyectosestudiantes`;